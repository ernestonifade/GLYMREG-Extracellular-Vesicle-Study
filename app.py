import os
import uuid
import openpyxl
import pandas as pd
import streamlit as st
from figures.figure1 import load_fig1_results, render_figure1
from figures.figure2 import load_fig2_results, render_figure2
from figures.figure3 import load_fig3_results, render_figure3
from figures.figure4 import load_fig4_results, render_figure4
from figures.figure5 import load_fig5_results, render_figure5
from figures.figure6 import load_fig6_results, render_figure6
from figures.figure7 import load_fig7_results, render_figure7
from figures.figure8 import load_fig8_results, render_figure8
from utils import render_searchable_table

# --- 1. STREAMLIT PAGE CONFIGURATION (WIDE & OPEN) ---
st.set_page_config(
    page_title="GLYMREG EV Study Dashboard",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
    .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        max-width: 98%;
    }
    h1 {
        font-size: 1.8rem !important;
        margin-bottom: 0px !important;
    }
    p {
        font-size: 0.95rem;
    }
    .stDownloadButton button {
        background-color: #28a745 !important;
        color: white !important;
        font-weight: bold;
        width: 100%;
    }
</style>
""",
    unsafe_allow_html=True,
)

st.title("🧬 GLYMREG Extracellular Vesicle Study")
st.caption("Interactive Manuscript Dashboard & Statistical Summary")
st.markdown("---")

# --- 2. SIDEBAR NAVIGATION ---
st.sidebar.header("📌 Navigation")
selected_figure = st.sidebar.radio(
    "Select Figure:",
    [
        "Figure 1: EV Size Skewness",
        "Figure 2: Extracellular Vesicles (Concentration, Size & Correlation)",
        "Figure 3: Proteomics (518 Panel)",
        "Figure 4: Cytokine Analysis",
        "Figure 5: Protein vs Cytokine vs Blood Correlation",
        "Figure 6: EV vs Protein, Cytokine vs Blood Correlations",
        "Figure 7: Pathway Enrichment Protein, Cytokine Correlations",
        "Figure 8: Biophysical predictors of EV concentration shifts",
    ],
    index=2,
)

st.sidebar.markdown("---")


# Helper function to write DataFrames safely using openpyxl
def export_sheets_to_excel(filename, sheets_dict):
  wb = openpyxl.Workbook()
  default_sheet = wb.active
  wb.remove(default_sheet)

  for sheet_name, df in sheets_dict.items():
    if df is not None and not df.empty:
      ws = wb.create_sheet(title=sheet_name)
      ws.append(list(df.columns))
      for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

  if len(wb.worksheets) == 0:
    ws = wb.create_sheet(title="Info")
    ws.append(["Note"])
    ws.append(["No data available"])

  wb.save(filename)


# --- 3. PAGE ROUTING & RENDER CALLS ---
if selected_figure == "Figure 1: EV Size Skewness":
  render_figure1()
elif (
    selected_figure
    == "Figure 2: Extracellular Vesicles (Concentration, Size & Correlation)"
):
  render_figure2()
elif selected_figure == "Figure 3: Proteomics (518 Panel)":
  render_figure3()
elif selected_figure == "Figure 4: Cytokine Analysis":
  render_figure4()
elif selected_figure == "Figure 5: Protein vs Cytokine vs Blood Correlation":
  render_figure5()
elif selected_figure == "Figure 6: EV vs Protein, Cytokine vs Blood Correlations":
  render_figure6()
elif (
    selected_figure
    == "Figure 7: Pathway Enrichment Protein, Cytokine Correlations"
):
  render_figure7()
elif selected_figure == "Figure 8: Biophysical predictors of EV concentration shifts":
  render_figure8()


# --- 4. ONE-CLICK INSTANT EXPORT HANDLER ---
st.sidebar.header("📥 Export Statistical Reports")

if selected_figure == "Figure 1: EV Size Skewness":
  stats_df, df_all = load_fig1_results()
  out_name = "Figure1D_EV_Size_Skewness_Data.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "EV_Size_Skewness_Wilcoxon_Stats": stats_df,
          "Raw_EV_Data": df_all,
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 1 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig1_{uuid.uuid4()}",
    )

elif (
    selected_figure
    == "Figure 2: Extracellular Vesicles (Concentration, Size & Correlation)"
):
  ancova_df, posthoc_df, corr_overall, corr_sex, long_df = load_fig2_results()
  out_name = "Figure2_EV_Full_Stats_Report.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "RM_ANCOVA_Stats": ancova_df,
          "PostHoc_Contrasts": posthoc_df,
          "Correlation_Overall": corr_overall,
          "Correlation_Sex": corr_sex,
          "Raw_Data": long_df,
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 2 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig2_{uuid.uuid4()}",
    )

elif selected_figure == "Figure 3: Proteomics (518 Panel)":
  ancova_df, posthoc_df, pca_scores_df, perm_df, long_df, _ = load_fig3_results()
  out_name = "Figure3_Proteomics_Full_Stats_Report.xlsx"
  sheets_data = {
      "RM_ANCOVA_Model_Stats": ancova_df,
      "PostHoc_Pairwise_Contrasts": posthoc_df,
      "PERMANOVA_Summary": perm_df,
      "PCA_Scores_and_EV": pca_scores_df,
      "Raw_Proteomic_Data": long_df,
  }
  for path in [
      "data/enrichment_permutation_results_for_interacting_proteins.csv",
      "enrichment_permutation_results_for_interacting_proteins.csv",
  ]:
    if os.path.exists(path):
      sheets_data["Protein_Interaction_Pathway_Enrichment"] = pd.read_csv(path)
      break

  export_sheets_to_excel(out_name, sheets_data)

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 3 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig3_{uuid.uuid4()}",
    )

elif selected_figure == "Figure 4: Cytokine Analysis":
  ancova_df, posthoc_df, df_emm, fig4_long_df = load_fig4_results()
  out_name = "Figure4_Cytokine_Full_Stats_Report.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "RM_ANCOVA_Model_Stats": ancova_df,
          "PostHoc_Pairwise_Contrasts": posthoc_df,
          "Group_EMM_Summary": df_emm,
          "Processed_Cytokine_Data": fig4_long_df,
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 4 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig4_{uuid.uuid4()}",
    )

elif selected_figure == "Figure 5: Protein vs Cytokine vs Blood Correlation":
  data = load_fig5_results()
  out_name = "Figure5_MultiModal_Integration_Report.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "Cyto_Protein_RM_Corr": data.get("cyto_rm"),
          "Blood_Protein_RM_Corr": data.get("blood_rm"),
          "Cyto_Protein_Baseline": data.get("cyto_baseline"),
          "Cyto_Protein_Delta_Windows": data.get("cyto_delta"),
          "Blood_Protein_Baseline": data.get("blood_baseline"),
          "Blood_Protein_Delta_Windows": data.get("blood_delta"),
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 5 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig5_{uuid.uuid4()}",
    )

elif (
    selected_figure == "Figure 6: EV vs Protein, Cytokine vs Blood Correlations"
):
  data = load_fig6_results()
  out_name = "Figure6_EV_MultiModal_Integration_Report.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "Cyto_Blood_RM_Corr": data.get("cyto_blood_rm"),
          "Protein_EVSize_RM_Corr": data.get("evsize_rm"),
          "Protein_EVConc_RM_Corr": data.get("evconc_rm"),
          "Cyto_Blood_Baseline": data.get("cyto_blood_baseline"),
          "Cyto_Blood_Delta_Windows": data.get("cyto_blood_delta"),
          "Protein_EVSize_Baseline": data.get("evsize_baseline"),
          "Protein_EVSize_Delta_Windows": data.get("evsize_delta"),
          "Protein_EVConc_Baseline": data.get("evconc_baseline"),
          "Protein_EVConc_Delta_Windows": data.get("evconc_delta"),
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 6 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig6_{uuid.uuid4()}",
    )

elif (
    selected_figure
    == "Figure 7: Pathway Enrichment Protein, Cytokine Correlations"
):
  def find_pathway_file(candidates):
    for path in candidates:
      if os.path.exists(path):
        return path
    return None

  out_name = "Figure7_Pathway_Enrichment_Report.xlsx"
  prot_path = find_pathway_file([
      "data/enrichment_permutation_results_for_correlating_proteins.csv",
      "../data/enrichment_permutation_results_for_correlating_proteins.csv",
      "enrichment_permutation_results_for_correlating_proteins.csv",
  ])
  cyt_path = find_pathway_file([
      "data/enrichment_permutation_results_for_correlating_cytokines.csv",
      "../data/enrichment_permutation_results_for_correlating_cytokines.csv",
      "enrichment_permutation_results_for_correlating_cytokines.csv",
  ])

  sheets_data = {}
  if prot_path and os.path.exists(prot_path):
    df_prot = pd.read_csv(prot_path)
    if not df_prot.empty:
      sheets_data["Protein_Pathways"] = df_prot

  if cyt_path and os.path.exists(cyt_path):
    df_cyt = pd.read_csv(cyt_path)
    if not df_cyt.empty:
      sheets_data["Cytokine_Pathways"] = df_cyt

  export_sheets_to_excel(out_name, sheets_data)

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 7 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig7_{uuid.uuid4()}",
    )

elif selected_figure == "Figure 8: Biophysical predictors of EV concentration shifts":
  pass
